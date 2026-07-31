"""
===============================================================================
Java EE 8 / Java 8  —  Method Extractor
===============================================================================

Extracts every method and constructor from a Java project and writes an Excel
file with:

    Column A  class               — simple class name
    Column B  class_method_key    — Class.method  (used as a unique key)
    Column C  Class.Method(args)  — full signature
    Column D  Number_Of_Lines     — LOC including annotations, excl. blank lines
    Column E  method_annotations  — e.g. @Override, @PostConstruct, @Schedule
    Column F  class_annotations   — e.g. @Stateless, @MessageDriven, @Singleton
    Column G  modifiers           — e.g. public, private, static
    Column H  entry_type          — EJB_SLSB | EJB_MDB | EJB_SINGLETON |
                                    SERVLET | FILTER | LISTENER | WEBSOCKET |
                                    JAX_RS | SCHEDULED | CDI | GENERAL
    Column I  parent_class        — extends X
    Column J  interfaces          — implements X, Y

Java EE 8 patterns handled
--------------------------
    @Stateless          EJB Stateless Session Bean
    @Stateful           EJB Stateful Session Bean
    @Singleton          EJB Singleton bean
    @MessageDriven      Message-Driven Bean (MDB) — JMS
    @WebServlet         HTTP Servlet
    @WebFilter          Servlet Filter
    @WebListener        ServletContextListener / HttpSessionListener
    @ServerEndpoint     WebSocket endpoint
    @Path               JAX-RS REST resource
    @ApplicationPath    JAX-RS Application
    @Schedule           Timer-driven method inside any EJB
    @PostConstruct      Lifecycle init method
    @PreDestroy         Lifecycle destroy method
    @Remote / @Local    EJB interface markers

Special method flags
--------------------
    onMessage()     — MDB entry point  (JMS)
    doGet/doPost()  — Servlet HTTP entry points
    init/destroy()  — Servlet lifecycle

Usage
-----
    from method_extractor import extract_methods_from_project
    extract_methods_from_project(
        extension=[".java"],
        project_path=r"C:\MyProject",
        output_path=r"C:\Output",
        excel_name="methods.xlsx"
    )

===============================================================================
"""

import os
import re

import javalang
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# =============================================================================
#  Java EE annotation → entry-type label
# =============================================================================

# Class-level annotations that define the EE component type
_CLASS_ENTRY_TYPE = {
    "Stateless"       : "EJB_SLSB",
    "Stateful"        : "EJB_SFSB",
    "Singleton"       : "EJB_SINGLETON",
    "MessageDriven"   : "EJB_MDB",
    "WebServlet"      : "SERVLET",
    "WebFilter"       : "FILTER",
    "WebListener"     : "LISTENER",
    "ServerEndpoint"  : "WEBSOCKET",
    "Path"            : "JAX_RS",
    "ApplicationPath" : "JAX_RS",
    "ApplicationScoped": "CDI",
    "RequestScoped"   : "CDI",
    "SessionScoped"   : "CDI",
    "Dependent"       : "CDI",
}

# Inherit-based entry type (no annotation needed)
_EXTENDS_ENTRY_TYPE = {
    "HttpServlet"   : "SERVLET",
    "GenericServlet": "SERVLET",
}

# Method-level annotations that are notable for Java EE
_NOTABLE_METHOD_ANNOTATIONS = {
    "PostConstruct", "PreDestroy", "Schedule", "Schedules",
    "TransactionAttribute", "RolesAllowed", "PermitAll", "DenyAll",
    "Asynchronous", "Lock", "Override",
    # JAX-RS method annotations
    "GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS",
    # WebSocket
    "OnOpen", "OnClose", "OnMessage", "OnError",
}

# MDB-specific entry method
_MDB_ENTRY_METHODS = {"onMessage"}

# Servlet lifecycle methods
_SERVLET_ENTRY_METHODS = {"doGet", "doPost", "doPut", "doDelete",
                           "doHead", "doOptions", "service",
                           "init", "destroy"}


# =============================================================================
#  LOC counter  (brace-matching, comment/string aware)
# =============================================================================

def get_method_line_count(
        details,
        java_folder,
        classname,
        methodname,
        *,
        java_file_path=None,
        line_cache=None,
        include_package_private=False,
        count_empty_lines=False,          # ← default False: skip blank lines
        parameter_signature=None,
        parameter_arity=None,
        parameter_types=None
):
    """
    Count lines of code for a specific method/constructor.

    Key Java EE 8 / Java 8 behaviours
    -----------------------------------
    • Skips blank lines by default (count_empty_lines=False)
    • Handles @Stateless / @MessageDriven / @Singleton class-level annotations
      (they appear above the class decl, not the method — correctly handled)
    • Works with javax.* namespace (Java 8 / EE 8), not jakarta.*
    • Overload-aware: picks correct overload by arity → type names → max LOC
    • Annotation block ABOVE the method signature is included in LOC
    • Abstract / interface methods without a body return LOC = 1
    """

    import os, re

    classname  = str(classname).strip()
    methodname = str(methodname).strip()

    cache_key = (
        (java_file_path or "").lower(),
        classname.lower(),
        methodname.lower(),
        include_package_private,
        count_empty_lines,
        str(parameter_arity),
        str(parameter_types),
    )
    if line_cache is not None and cache_key in line_cache:
        return line_cache[cache_key]

    # ── locate file ──────────────────────────────────────────────────
    if not java_file_path:
        target = f"{classname}.java".lower()
        for root, _, files in os.walk(java_folder):
            for f in files:
                if f.lower() == target:
                    java_file_path = os.path.join(root, f)
                    break
            if java_file_path:
                break

    if not java_file_path:
        if line_cache is not None:
            line_cache[cache_key] = None
        return None

    # ── read ─────────────────────────────────────────────────────────
    try:
        text = open(java_file_path, encoding="utf-8").read()
    except Exception:
        try:
            text = open(java_file_path, encoding="latin-1").read()
        except Exception:
            return None

    text  = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")

    # ── comment/string-aware helpers ─────────────────────────────────

    def _scan(text, start, look_for):
        """
        Generic forward scanner that respects // /* */ string and char literals.
        look_for: callable(ch, nxt, nxt2) → returns position if found, else None
        """
        in_block = in_line = in_str = False
        str_ch = None
        i = start
        n = len(text)
        while i < n:
            ch   = text[i]
            nxt  = text[i+1] if i+1 < n else ""
            nxt2 = text[i+2] if i+2 < n else ""

            if in_block:
                if ch == "*" and nxt == "/": in_block = False; i += 2; continue
                i += 1; continue
            if in_line:
                if ch == "\n": in_line = False
                i += 1; continue
            if in_str:
                if ch == "\\": i += 2; continue
                if ch == str_ch: in_str = False; str_ch = None
                i += 1; continue

            if ch == "/" and nxt == "*": in_block = True; i += 2; continue
            if ch == "/" and nxt == "/": in_line = True;  i += 2; continue
            if ch in ('"', "'"): in_str = True; str_ch = ch; i += 1; continue

            result = look_for(ch, nxt, nxt2, i)
            if result is not None:
                return result
            i += 1
        return None

    def find_open_brace(start):
        def _check_open(ch, nxt, nxt2, pos):
            return pos if ch == "{" else None
        return _scan(text, start, _check_open)

    def find_matching_close(open_pos):
        depth = [0]
        def check(ch, nxt, nxt2, pos):
            if ch == "{": depth[0] += 1
            elif ch == "}":
                depth[0] -= 1
                if depth[0] == 0: return pos
            return None
        return _scan(text, open_pos, check)

    def find_matching_paren(open_pos):
        depth = [0]
        angle = [0]
        def check(ch, nxt, nxt2, pos):
            if ch == "<": angle[0] += 1
            elif ch == ">" and angle[0] > 0: angle[0] -= 1
            elif ch == "(": depth[0] += 1
            elif ch == ")":
                depth[0] -= 1
                if depth[0] == 0: return pos
            return None
        return _scan(text, open_pos, check)

    # ── locate class declaration ──────────────────────────────────────
    class_kw = r"(?:class|interface|enum|record)"
    cls_re = re.compile(
        rf"(?m)^[ \t]*(?:@\w+(?:\([^)]*\))?[ \t\n]*)*"
        rf"(?:public|protected|private)?[ \t]*"
        rf"(?:(?:abstract|final|static|strictfp|sealed|non-sealed)[ \t]+)*"
        rf"{class_kw}[ \t]+{re.escape(classname)}\b"
    )
    cm = cls_re.search(text)
    if not cm:
        # fallback: unanchored
        cls_re2 = re.compile(
            rf"(?:class|interface|enum|record)[ \t]+{re.escape(classname)}\b"
        )
        cm = cls_re2.search(text)
    if not cm:
        if line_cache is not None: line_cache[cache_key] = None
        return None

    cls_open  = text.find("{", cm.end())
    if cls_open == -1:
        if line_cache is not None: line_cache[cache_key] = 1
        return 1
    cls_close = find_matching_close(cls_open)
    if cls_close is None: cls_close = len(text) - 1

    class_block        = text[cls_open : cls_close + 1]
    cls_block_offset   = cls_open
    cls_block_start_ln = text.count("\n", 0, cls_open) + 1

    # ── method signature patterns ─────────────────────────────────────
    access_opt  = r"(?:(?:public|private|protected)\s+)?"
    access_req  = r"(?:public|private|protected)\s+"
    access      = access_opt if include_package_private else access_req
    mods        = r"(?:(?:static|final|abstract|synchronized|native|strictfp|default)\s+)*"
    mn          = re.escape(methodname)

    method_re = re.compile(
        rf"(?m)^[ \t]*{access}{mods}"
        rf"(?:<[^>]*>\s*)?"
        rf"[A-Za-z_][\w.<>\[\],\s?]*\s+"
        rf"{mn}[ \t]*\(",
        re.IGNORECASE,
    )
    ctor_re = re.compile(
        rf"(?m)^[ \t]*{access}{mods}"
        rf"\b{re.escape(classname)}[ \t]*\(",
        re.IGNORECASE,
    )

    matches = (
        list(ctor_re.finditer(class_block))
        if methodname == classname
        else list(method_re.finditer(class_block))
    )

    if not matches:
        # Java EE private lifecycle methods (@PostConstruct, @Schedule etc.)
        # allow package-private / private — retry without access requirement
        method_re_pp = re.compile(
            rf"(?m)^[ \t]*(?:(?:public|private|protected|static|final|abstract"
            rf"|synchronized|native|strictfp|default)\s+)*"
            rf"(?:<[^>]*>\s*)?"
            rf"[A-Za-z_][\w.<>\[\],\s?]*\s+"
            rf"{mn}[ \t]*\(",
            re.IGNORECASE,
        )
        matches = list(method_re_pp.finditer(class_block))

    if not matches:
        if line_cache is not None: line_cache[cache_key] = None
        return None

    # ── arity / type helpers ──────────────────────────────────────────
    def parse_params(region):
        s = re.sub(r'@\w+(?:\([^)]*\))?', '', region)
        s = re.sub(r'<[^>]*>', '', s).replace("\n", " ")
        parts, buf, depth = [], "", 0
        for ch in s:
            if ch == "(": depth += 1; buf += ch
            elif ch == ")": depth = max(0, depth-1); buf += ch
            elif ch == "," and depth == 0: parts.append(buf.strip()); buf = ""
            else: buf += ch
        if buf.strip(): parts.append(buf.strip())
        if len(parts) == 1 and parts[0] == "":
            return 0, []
        types = []
        for p in parts:
            p = p.split("=", 1)[0].strip().replace("...", "[]")
            p = re.sub(r'\b(final|volatile|transient)\b', '', p)
            toks = re.findall(r'[A-Za-z_]\w+|\[\]', p)
            if not toks: types.append(""); continue
            arr = ""
            while toks and toks[-1] == "[]": arr += "[]"; toks.pop()
            if not toks: types.append(arr); continue
            toks.pop()   # drop param name
            t = next((x for x in reversed(toks) if x != "[]"), "")
            types.append(t + arr)
        return len(parts), types

    def types_equal(a, b):
        def n(x):
            x = (x or "").strip().split(".")[-1]
            x = re.sub(r'\[\]+$', '[]', x)
            return x.lower()
        return n(a) == n(b)

    # ── compute LOC for one match ─────────────────────────────────────
    def loc_for(m):
        sig_abs_start = cls_block_offset + m.start()
        sig_abs_end   = cls_block_offset + m.end()
        sig_line      = text.count("\n", 0, sig_abs_start) + 1

        # walk upward to include contiguous annotation lines
        def anno_start(sig_ln):
            i = sig_ln - 2
            pbal, started, start_ln = 0, False, None
            while i >= cls_block_start_ln - 1:
                ln = lines[i].rstrip()
                if not ln.strip() and not (started and pbal > 0): break
                is_ann = ln.lstrip().startswith("@")
                if not started:
                    if is_ann:
                        started = True; start_ln = i + 1
                        pbal = ln.count("(") - ln.count(")")
                    else: break
                else:
                    if is_ann or pbal > 0:
                        start_ln = i + 1
                        pbal += ln.count("(") - ln.count(")")
                    else: break
                i -= 1
            return start_ln

        start_ln = anno_start(sig_line) or sig_line

        brace_open = find_open_brace(sig_abs_end)
        if brace_open is None:
            return 1   # abstract / interface method

        brace_close = find_matching_close(brace_open)
        if brace_close is None: brace_close = len(text) - 1

        end_ln = text.count("\n", 0, brace_close) + 1

        if count_empty_lines:
            return max(1, end_ln - start_ln + 1)
        segment = lines[start_ln - 1 : end_ln]
        return max(1, sum(1 for l in segment if l.strip()))

    # ── collect candidates ────────────────────────────────────────────
    candidates = []
    for m in matches:
        paren_open = cls_block_offset + m.end() - 1
        paren_close = find_matching_paren(paren_open)
        if paren_close is None:
            candidates.append({"arity": None, "types": [], "loc": loc_for(m)})
            continue
        arity, types = parse_params(text[paren_open+1 : paren_close])
        candidates.append({"arity": arity, "types": types, "loc": loc_for(m)})

    # ── pick best overload ────────────────────────────────────────────
    target_arity = None
    if parameter_arity is not None:
        try: target_arity = int(parameter_arity)
        except: pass

    target_types = [t.strip() for t in str(parameter_types or "").split(";")
                    if t and t.strip()]

    pool = candidates
    if target_arity is not None:
        pool = [c for c in pool if c["arity"] == target_arity] or pool

    if len(pool) > 1 and target_types:
        def score(c):
            if len(c["types"]) != len(target_types): return -1
            return sum(1 for i, tt in enumerate(target_types)
                       if types_equal(c["types"][i], tt))
        max_s = max(score(c) for c in pool)
        pool = [c for c in pool if score(c) == max_s]

    best = max(c["loc"] for c in pool) if pool else None
    if line_cache is not None:
        line_cache[cache_key] = best
    return best


# =============================================================================
#  Java EE entry-type classifier
# =============================================================================

def classify_entry_type(class_annotations: list, extends: str,
                        interfaces: list) -> str:
    """
    Return the Java EE component label for a class based on its annotations
    and inheritance.  Falls back to GENERAL for plain Java classes.
    """
    for ann in class_annotations:
        if ann in _CLASS_ENTRY_TYPE:
            return _CLASS_ENTRY_TYPE[ann]

    # inheritance-based (e.g. extends HttpServlet without @WebServlet)
    if extends and extends in _EXTENDS_ENTRY_TYPE:
        return _EXTENDS_ENTRY_TYPE[extends]

    # implements MessageListener → likely MDB
    if interfaces and "MessageListener" in interfaces:
        return "EJB_MDB"

    return "GENERAL"


# =============================================================================
#  Method extractor  (javalang AST — works perfectly for Java 8 / EE 8)
# =============================================================================

def extract_methods_from_file(file_path: str) -> list:
    """
    Parse a .java file with javalang and return a list of method dicts.

    Each dict contains:
        class, method, full_args,
        parent_class, interfaces,
        class_annotations, method_annotations,
        modifiers,
        java_file_path
    """
    try:
        src = open(file_path, encoding="utf-8", errors="ignore").read()
        tree = javalang.parse.parse(src)
    except Exception as e:
        print(f"  ⚠ Parse error — skipping: {os.path.basename(file_path)} ({e})")
        return []

    results = []

    for _path, node in tree.filter(javalang.tree.TypeDeclaration):
        class_name = node.name

        # class-level annotations (e.g. @Stateless, @MessageDriven)
        class_annotations = [a.name for a in (node.annotations or [])]

        # parent class
        parent = getattr(node, "extends", None)
        if parent is None:
            parent_name = None
        elif isinstance(parent, list):
            parent_name = ", ".join(
                p.name if hasattr(p, "name") else str(p) for p in parent
            )
        else:
            parent_name = parent.name if hasattr(parent, "name") else str(parent)

        # interfaces
        ifaces_raw = getattr(node, "implements", []) or []
        interfaces = [
            i.name if hasattr(i, "name") else str(i) for i in ifaces_raw
        ]

        def _args(params):
            args = []
            for p in params:
                tn = (
                    p.type.name
                    if getattr(p.type, "name", None)
                    else str(p.type)
                )
                # handle array dimensions
                dims = getattr(p.type, "dimensions", None)
                if dims:
                    tn += "[]" * len(dims)
                args.append(f"{tn} {p.name}")
            return ", ".join(args)

        # ── constructors ──────────────────────────────────────────────
        for ctor in getattr(node, "constructors", []):
            mods = list(ctor.modifiers or [])
            anns = [a.name for a in (ctor.annotations or [])]
            results.append({
                "class"             : class_name,
                "method"            : ctor.name,
                "full_args"         : _args(ctor.parameters),
                "parent_class"      : parent_name,
                "interfaces"        : interfaces,
                "class_annotations" : class_annotations,
                "method_annotations": anns,
                "modifiers"         : mods,
                "java_file_path"    : file_path,
            })

        # ── methods ───────────────────────────────────────────────────
        for method in getattr(node, "methods", []):
            mods = list(method.modifiers or [])
            anns = [a.name for a in (method.annotations or [])]
            results.append({
                "class"             : class_name,
                "method"            : method.name,
                "full_args"         : _args(method.parameters),
                "parent_class"      : parent_name,
                "interfaces"        : interfaces,
                "class_annotations" : class_annotations,
                "method_annotations": anns,
                "modifiers"         : mods,
                "java_file_path"    : file_path,
            })

    return results


# =============================================================================
#  Excel writer
# =============================================================================

# Colour palette
_HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue header
_EJB_FILL   = PatternFill("solid", fgColor="D6E4F0")   # light blue  — EJB
_SERVLET_FILL = PatternFill("solid", fgColor="E2EFDA") # light green — servlet
_MDB_FILL   = PatternFill("solid", fgColor="FCE4D6")   # light orange— MDB
_CDI_FILL   = PatternFill("solid", fgColor="FFF2CC")   # light yellow— CDI
_GENERAL_FILL = PatternFill("solid", fgColor="F2F2F2") # light grey  — general

_ROW_FILLS = {
    "EJB_SLSB"     : _EJB_FILL,
    "EJB_SFSB"     : _EJB_FILL,
    "EJB_SINGLETON": _EJB_FILL,
    "EJB_MDB"      : _MDB_FILL,
    "SERVLET"      : _SERVLET_FILL,
    "FILTER"       : _SERVLET_FILL,
    "LISTENER"     : _SERVLET_FILL,
    "WEBSOCKET"    : _SERVLET_FILL,
    "JAX_RS"       : _CDI_FILL,
    "CDI"          : _CDI_FILL,
    "SCHEDULED"    : _EJB_FILL,
    "GENERAL"      : _GENERAL_FILL,
}

_HEADERS = [
    "class",
    "class_method_key",
    "Class.Method",
    "Number_Of_Lines",
    "method_annotations",
    "class_annotations",
    "modifiers",
    "entry_type",
    "parent_class",
    "interfaces",
]


def write_to_excel(methods_data: list, project_path: str,
                   output_excel: str) -> str:

    wb = Workbook()
    ws = wb.active
    ws.title = "Methods"

    # ── header row ────────────────────────────────────────────────────
    for col, h in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[1].height = 25

    # ── data rows ─────────────────────────────────────────────────────
    line_cache = {}

    for row_idx, m in enumerate(methods_data, start=2):

        class_anns  = m.get("class_annotations", [])
        method_anns = m.get("method_annotations", [])
        modifiers   = m.get("modifiers", [])
        parent      = m.get("parent_class") or ""
        ifaces      = m.get("interfaces", [])
        entry_type  = classify_entry_type(
            class_anns,
            parent,
            ifaces,
        )

        # demote @Schedule methods to SCHEDULED even inside a SINGLETON
        if "Schedule" in method_anns or "Schedules" in method_anns:
            entry_type = "SCHEDULED"

        loc = get_method_line_count(
            details          = {},
            java_folder      = project_path,
            classname        = m["class"],
            methodname       = m["method"],
            java_file_path   = m["java_file_path"],
            line_cache       = line_cache,
            parameter_signature = m["full_args"],
            count_empty_lines   = False,
        )

        row_data = [
            m["class"],
            f"{m['class']}.{m['method']}",
            f"{m['class']}.{m['method']}({m['full_args']})",
            loc,
            ", ".join(f"@{a}" for a in method_anns) if method_anns else "",
            ", ".join(f"@{a}" for a in class_anns)  if class_anns  else "",
            ", ".join(modifiers) if modifiers else "",
            entry_type,
            parent,
            ", ".join(ifaces) if ifaces else "",
        ]

        fill = _ROW_FILLS.get(entry_type, _GENERAL_FILL)

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(wrap_text=False)

    # ── column widths ─────────────────────────────────────────────────
    col_widths = [22, 35, 55, 16, 35, 35, 20, 16, 22, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[
            ws.cell(row=1, column=i).column_letter
        ].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_excel)
    print(f"\n  ✓ Excel written: {output_excel}")
    print(f"    Rows: {len(methods_data)}   Columns: {len(_HEADERS)}")
    return output_excel


# =============================================================================
#  Public API
# =============================================================================

def extract_methods_from_project(
        extension: list,
        project_path: str,
        output_path: str,
        excel_name: str,
) -> str:
    """
    Walk project_path, extract every method from files matching extension,
    and write an Excel report to output_path/excel_name.

    Parameters
    ----------
    extension    : list of file suffixes, e.g. [".java"]
    project_path : root of the Java / Java EE project
    output_path  : directory where the Excel file is written
    excel_name   : filename, e.g. "methods.xlsx"

    Returns
    -------
    Full path to the written Excel file.
    """

    print(f"\nScanning: {project_path}")

    ext_tuple = tuple(extension)
    all_methods = []
    file_count  = 0

    for root, dirs, files in os.walk(project_path):
        # skip build output folders
        dirs[:] = [d for d in dirs if d not in
                   {"target", "bin", "build", "out", ".git", "node_modules"}]
        for fname in files:
            if fname.endswith(ext_tuple):
                fp = os.path.join(root, fname)
                methods = extract_methods_from_file(fp)
                all_methods.extend(methods)
                file_count += 1

    print(f"  Files scanned : {file_count}")
    print(f"  Methods found : {len(all_methods)}")

    os.makedirs(output_path, exist_ok=True)
    output_excel = os.path.join(output_path, excel_name)

    return write_to_excel(all_methods, project_path, output_excel)


# =============================================================================
#  Quick self-test when run directly
# =============================================================================
if __name__ == "__main__":

    PROJECT_PATH = r"C:\Users\Barath\Downloads\sample.daytrader7-master_new"
    OUTPUT_PATH  = r"C:\Users\Barath\Downloads\sample.daytrader7-master_new"
    EXCEL_NAME   = "methods.xlsx"

    extract_methods_from_project(
        extension    = [".java"],
        project_path = PROJECT_PATH,
        output_path  = OUTPUT_PATH,
        excel_name   = EXCEL_NAME,
    )