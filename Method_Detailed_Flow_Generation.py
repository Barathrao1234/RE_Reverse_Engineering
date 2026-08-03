import pandas as pd
import numpy as np
import os
import re
from collections import defaultdict
import logging
from dotenv import load_dotenv
load_dotenv()
# ---------------- CONFIG ----------------
CASE_INSENSITIVE = True
MERGE_CELLS = True
MAX_DEPTH = None  # Optional depth limit


def get_method_line_count(
    details,
    java_folder,
    classname,
    methodname,
    line_cache=None,
    include_package_private=False,
    count_empty_lines=True,  # <<< FLAG: True = count blanks, False = skip blanks
):
    
    # ---------- Cache ----------
    if line_cache is not None:
        cache_key = (
            classname.lower(),
            methodname.lower(),
            include_package_private,
            "auto_anno_inclusive",
            count_empty_lines,  # include flag in cache key
        )
        if cache_key in line_cache:
            return line_cache[cache_key]

    # ---------- Locate Java source ----------
    target_filename = f"{classname}{details.get('extension', '.java')}".lower()
    java_file_path = None
    for root, _, files in os.walk(java_folder):
        for file in files:
            if file.lower() == target_filename:
                java_file_path = os.path.join(root, file)
                break
        if java_file_path:
            break
    if not java_file_path:
        return None

    # ---------- Read file ----------
    try:
        with open(java_file_path, "r", encoding="utf-8") as f:
            text = f.read()
    except Exception:
        try:
            with open(java_file_path, "r", encoding="latin-1") as f:
                text = f.read()
        except Exception:
            return None

    # Normalize EOLs
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")

    # ---------- Regex: find the METHOD DECLARATION line (not annotations) ----------
    methodname_esc = re.escape(methodname)
    access_req = r"(?:public|private|protected)"
    access = rf"(?:{access_req})?" if include_package_private else access_req

    method_decl_regex = re.compile(
        rf"""
        (?m)                                   # multiline
        ^[ \t]*                                # start of the declaration line
        {access}[ \t]*                         # access (optional if include_package_private=True)
        (?:(?:static|final|abstract|synchronized|native|strictfp)\b[ \t]*)*  # optional modifiers
        [\w.<>\[\],? \t]+                      # return type (permissive)
        \b(?P<mname>{methodname_esc})[ \t]*\(
        """,
        re.IGNORECASE | re.VERBOSE,
    )

    m = method_decl_regex.search(text)
    if not m:
        return None

    # Signature line index (1-based)
    sig_line_idx = text.count("\n", 0, m.start("mname")) + 1

    # ---------- Upward scan to include immediate annotation block ----------
    def find_annotation_block_start(signature_line_index: int):
        i = signature_line_index - 2  # line above signature, 0-based
        if i < 0:
            return None

        paren_balance = 0
        started = False
        start_line = None

        while i >= 0:
            line = lines[i].rstrip()

            # Stop if blank and not inside a continued annotation
            if not line.strip() and not (started and paren_balance > 0):
                break

            is_anno_start = bool(re.match(r'^[ \t]*@', line))

            if not started:
                if is_anno_start:
                    started = True
                    start_line = i + 1  # 1-based
                    paren_balance = line.count("(") - line.count(")")
                else:
                    break
            else:
                if is_anno_start or paren_balance > 0:
                    start_line = i + 1
                    paren_balance += line.count("(") - line.count(")")
                else:
                    break

            i -= 1

        return start_line

    anno_start_idx = find_annotation_block_start(sig_line_idx)
    start_line_idx = anno_start_idx if anno_start_idx is not None else sig_line_idx

    # ---------- Find opening brace line (IGNORE braces in annotations) ----------
    # IMPORTANT CHANGE: start scanning from the SIGNATURE line, not from annotation start.
    def find_opening_brace_line(start_from_line_index: int):
        in_block_comment = False
        for i in range(start_from_line_index - 1, len(lines)):  # 0-based
            line = lines[i]
            j, n = 0, len(line)
            in_string = False
            string_char = None
            while j < n:
                ch = line[j]
                nxt = line[j + 1] if j + 1 < n else ""

                if in_block_comment:
                    if ch == "*" and nxt == "/":
                        in_block_comment = False
                        j += 2
                        continue
                    j += 1
                    continue

                if in_string:
                    if ch == "\\":
                        j += 2
                        continue
                    if ch == string_char:
                        in_string = False
                        string_char = None
                    j += 1
                    continue

                if ch == "/" and nxt == "*":
                    in_block_comment = True
                    j += 2
                    continue
                if ch == "/" and nxt == "/":
                    break  # ignore rest of line

                if ch in ("'", '"'):
                    in_string = True
                    string_char = ch
                    j += 1
                    continue

                if ch == "{":
                    return i + 1  # 1-based

                j += 1
        return None

    # <<< HERE: Start from sig_line_idx to avoid '{' in annotations >>>
    brace_start_line = find_opening_brace_line(sig_line_idx)
    if brace_start_line is None:
        # No body; count at least the signature/annotation line
        result = 1 if count_empty_lines else (1 if lines[start_line_idx - 1].strip() else 0) or 1
        if line_cache is not None:
            line_cache[cache_key] = result
        return result

    # ---------- Find closing brace line by tracking depth (ignore braces in strings/comments) ----------
    def find_closing_brace_line(open_line_index: int):
        in_block_comment = False
        depth = 0
        started = False
        for i in range(open_line_index - 1, len(lines)):  # 0-based
            line = lines[i]
            j, n = 0, len(line)
            in_string = False
            string_char = None
            while j < n:
                ch = line[j]
                nxt = line[j + 1] if j + 1 < n else ""

                if in_block_comment:
                    if ch == "*" and nxt == "/":
                        in_block_comment = False
                        j += 2
                        continue
                    j += 1
                    continue

                if in_string:
                    if ch == "\\":
                        j += 2
                        continue
                    if ch == string_char:
                        in_string = False
                        string_char = None
                    j += 1
                    continue

                if ch == "/" and nxt == "*":
                    in_block_comment = True
                    j += 2
                    continue

                if ch == "/" and nxt == "/":
                    break  # rest is line-comment

                if ch in ("'", '"'):
                    in_string = True
                    string_char = ch
                    j += 1
                    continue

                if ch == "{":
                    depth += 1
                    started = True
                elif ch == "}":
                    depth -= 1
                    if started and depth == 0:
                        return i + 1  # 1-based
                j += 1
        return None

    end_idx = find_closing_brace_line(brace_start_line)
    if end_idx is None:
        end_idx = len(lines)  # fallback if unbalanced

    # ---------- Final count ----------
    if count_empty_lines:
        line_count = max(1, end_idx - start_line_idx + 1)
    else:
        segment = lines[start_line_idx - 1 : end_idx]
        non_empty = sum(1 for ln in segment if ln.strip() != "")
        line_count = max(1, non_empty)

    if line_cache is not None:
        line_cache[cache_key] = line_count

    return line_count
def precompute_line_counts(details,java_folder, df_orig):
    line_cache = {}
    unique_methods = df_orig[['classname', 'methodname']].drop_duplicates()
    
    for i, (_, row) in enumerate(unique_methods.iterrows()):
        
        key = (row['classname'].lower(), row['methodname'].lower())
        result = get_method_line_count(details,java_folder, row['classname'], row['methodname'])
        line_cache[key] = result
    return line_cache

# ---------------- METHOD: Read AST & Build Graph ----------------
# ---------------- METHOD: Read AST & Build Graph (with _row_order preserved) ----------------
def read_ast_and_build_graph(excel_path, sheet_name="Cleaned_AST_Details"):
    df_orig = pd.read_excel(excel_path, sheet_name=sheet_name, engine="openpyxl")
    df_orig = df_orig.fillna("").astype(str)

    # Add row order immediately
    if '_row_order' not in df_orig.columns:
        df_orig['_row_order'] = np.arange(len(df_orig))

    required = ["file_name", "class_interface_name", "type", "method_name", "Annotations",
                "Method_Declaration_Type", "return_type", "object_call", "class_method_call"]
    missing = [c for c in required if c not in df_orig.columns]
    if missing:
        raise ValueError(f"Missing column(s): {missing}")

    # Keep _row_order while filtering required columns
    df_orig = df_orig[required + ['_row_order']].copy()

    # Rename columns but keep _row_order
    df_orig.rename(columns={
        "file_name": "filename",
        "class_interface_name": "classname",
        "method_name": "methodname"
    }, inplace=True)

    # Build maps for graph
    class_method_map = defaultdict(set)
    method_map = defaultdict(set)
    for _, row in df_orig.iterrows():
        cm_call = row["class_method_call"]
        if cm_call and cm_call.lower() not in ("none", "null", "na", ""):
            class_method_map[(row["classname"], row["methodname"])].add(cm_call.strip())
            method_map[row["methodname"]].add(cm_call.strip())

    # Convert sets to lists
    class_method_map = {k: list(v) for k, v in class_method_map.items()}
    method_map = {k: list(v) for k, v in method_map.items()}

    graph = defaultdict(list)
    method_to_nodes = defaultdict(list)
    filebase_to_nodes = defaultdict(list)

    for _, r in df_orig.iterrows():
        filebase = os.path.splitext(r["filename"])[0]
        node = f"{filebase}.{r['methodname']}"
        method_to_nodes[r["methodname"]].append(node)
        filebase_to_nodes[filebase].append(node)
        if node not in graph:
            graph[node] = []

    for _, r in df_orig.iterrows():
        caller = f"{os.path.splitext(r['filename'])[0]}.{r['methodname']}"
        raw = r["class_method_call"]
        if raw.lower() in ("", "none", "null", "na"):
            continue
        callee = raw.split("(")[0].strip()
        targets = []
        if "." in callee:
            targets.append(callee)
        else:
            targets.extend(method_to_nodes.get(callee, []))
        if not targets:
            targets.append(callee)
        for t in targets:
            if t not in graph:
                graph[t] = []
            graph[caller].append(t)

    for k in graph:
        graph[k] = list(dict.fromkeys(graph[k]))

    method_to_row = {}
    for _, row in df_orig.iterrows():
        key = (row['classname'].lower(), row['methodname'].lower())
        if key not in method_to_row:
            method_to_row[key] = row

    return df_orig, class_method_map, method_map, graph, filebase_to_nodes, method_to_row

# ---------------- PRECOMPUTE METHODNAME TO ROWS ----------------
def build_methodname_lookup(df_orig):
    methodname_to_rows = defaultdict(list)
    for _, row in df_orig.iterrows():
        methodname_to_rows[row['methodname'].lower()].append(row)
    return methodname_to_rows

# ---------------- RESOLVE UNQUALIFIED METHOD WITH CACHE ----------------
def resolve_unqualified_method(unqualified_method, df_orig, caller_row, methodname_to_rows, resolved_cache):
    key = unqualified_method.lower()
    if key in resolved_cache:
        return resolved_cache[key]

    qualified = []

    caller_class_interface = str(caller_row['classname']).strip()
    caller_object_call = str(caller_row['object_call']).strip()

    if "." in caller_object_call:
        caller_obj_class, caller_obj_method = caller_object_call.split(".", 1)
    else:
        caller_obj_class, caller_obj_method = caller_class_interface, caller_object_call

    rows = methodname_to_rows.get(key, [])

    for row in rows:
        type_val = str(row['type']).lower()
        access_val = str(row['Method_Declaration_Type']).lower()
        annotations = str(row['Annotations']).strip()
        return_type = str(row['return_type']).lower()

        candidate_class = str(row['classname']).strip()
        candidate_method = str(row['methodname']).strip()

        if type_val == "class" and access_val == "private":
            if f"{caller_obj_class}.{caller_obj_method}".lower() != f"{candidate_class}.{candidate_method}".lower():
                continue
            qualified.append(f"{candidate_class}.{candidate_method}")
            continue

        if type_val == "class" and access_val == "public":
            qualified.append(f"{candidate_class}.{candidate_method}")
            continue

        if type_val == "class_implements_interface" and "@Override" in annotations:
            qualified.append(f"{candidate_class}.{candidate_method}")
            continue

        if type_val == "interface":
            qualified.append(f"{candidate_class}.{candidate_method}")
            continue

    qualified = list(dict.fromkeys(qualified))
    resolved_cache[key] = qualified
    return qualified

# ---------------- EXPAND LINEAGE HORIZONTAL ----------------
def expand_lineage_horizontal(
    classname, methodname,
    class_method_map, method_map,
    visited, method_line_map,
    df_orig, method_to_row, methodname_to_rows,
    resolved_cache, dvt_set, depth=0
):
    # Skip DVT files
    if classname.lower() in dvt_set:
        return []

    key = (classname.lower(), methodname.lower())
    orig_row = method_to_row.get(key)

    # ---------- External / unresolved method ----------
    if orig_row is None:
        ext_line_count = method_line_map.get(
            (classname.lower(), methodname.lower()), 0
        )
        external_node = f"{classname}.{methodname} no_of_lines : {ext_line_count}"
        return [[external_node]]

    # ---------- Build node ----------
    line_count = method_line_map.get(
        (classname.lower(), methodname.lower()), 0
    )
    node_name = f"{classname}.{methodname} no_of_lines : {line_count}"

    # Max depth guard
    if MAX_DEPTH and depth >= MAX_DEPTH:
        return [[node_name]]

    # Cycle guard
    if node_name in visited:
        return [[node_name]]

    visited.add(node_name)

    children = class_method_map.get((classname, methodname), [])

    # Leaf node
    if not children:
        visited.remove(node_name)
        return [[node_name]]

    all_paths = []

    for child in children:
        sub_children = []

        # ---------- Resolve unqualified methods ----------
        if "." not in child:
            qualified_calls = resolve_unqualified_method(
                child, df_orig, orig_row,
                methodname_to_rows, resolved_cache
            )
            if qualified_calls:
                sub_children.extend(qualified_calls)
            else:
                sub_children.append(f"{classname}.{child}")
        else:
            sub_children.append(child)

        # ---------- Recurse ----------
        for subchild in sub_children:
            if "." in subchild:
                cls, meth = subchild.rsplit(".", 1)

                impl_cls = f"{cls}Impl"

                impl_key = (impl_cls.lower(), meth.lower())

                if impl_key in method_to_row:
                    cls = impl_cls

            else:
                cls, meth = classname, subchild
            sub_paths = expand_lineage_horizontal(
                cls, meth,
                class_method_map, method_map,
                visited, method_line_map,
                df_orig, method_to_row,
                methodname_to_rows, resolved_cache,
                dvt_set, depth + 1
            )

            for sp in sub_paths:
                all_paths.append([node_name] + sp)

    visited.remove(node_name)

    # Safety fallback
    if not all_paths:
        return [[node_name]]

    return all_paths

def flatten_list(lst):
    result = []
    for item in lst:
        if isinstance(item, list):
            result.extend(flatten_list(item))
        else:
            result.append(item)
    return result

# ---------------- METHOD: Build DataFrame ----------------
def build_level_dataframe_from_paths(paths):
    if not paths:
        return pd.DataFrame()
    max_len = max(len(p) for p in paths)
    cols = [f"Level_{i}" for i in range(1, max_len + 1)]
    rows = [p + [""] * (max_len - len(p)) for p in paths]
    return pd.DataFrame(rows, columns=cols)

# ---------------- METHOD: Compute occurrences ----------------
def compute_occurrences_from_df_levels(df_levels, java_roots, ast_df):
    file_bases = sorted({os.path.splitext(s)[0] for s in ast_df['filename'].astype(str).unique()})
    results = []
    if df_levels.empty:
        return pd.DataFrame(results, columns=["file_name", "Java_file", "Level", "contains_count"])

    # df_norm = df_levels.fillna("").astype(str).map(lambda s: s.lower().strip())
    df_temp = df_levels.fillna("").astype(str)

    if hasattr(df_temp, "map"):
        df_norm = df_temp.map(lambda s: s.lower().strip())
    else:
        df_norm = df_temp.applymap(lambda s: s.lower().strip())
        
    for fb in file_bases:
        fb_l = fb.lower()
        for col_idx, col in enumerate(df_norm.columns):
            for row_idx, cell in enumerate(df_norm[col]):
                if fb_l in cell:
                    results.append({
                        "file_name": fb,
                        "Java_file": java_roots[row_idx] if row_idx < len(java_roots) else "<separator>",
                        "Level": col,
                        "contains_count": 1
                    })
    return pd.DataFrame(results, columns=["file_name", "Java_file", "Level", "contains_count"])

# ---------------- METHOD: Write Excel ----------------
def write_all_sheets_pandas(METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL, df_levels, df_inverted_levels, occ_df):
    import xlsxwriter

    df_levels = df_levels.copy()
    df_levels['_row_order'] = np.arange(len(df_levels))

    with pd.ExcelWriter(METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL, engine='xlsxwriter') as writer:
        # Write original flow
        df_levels.drop(columns=['_row_order']).to_excel(writer, sheet_name="Original Flow", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Original Flow"]

        merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})

        data = df_levels.drop(columns=['_row_order']).to_numpy()
        n_rows, n_cols = data.shape

        for col_idx in range(n_cols):
            col_data = data[:, col_idx]
            change_idx = np.flatnonzero(np.r_[True, col_data[1:] != col_data[:-1], True])
            for start, end in zip(change_idx[:-1], change_idx[1:]):
                if end - start > 1:
                    worksheet.merge_range(start + 1, col_idx, end, col_idx, col_data[start], merge_format)

        df_inverted_levels.to_excel(writer, sheet_name="Inverted Flow", index=False)

        # if occ_df is not None and not occ_df.empty:
        #     occ_df.to_excel(writer, sheet_name="FileNameOccurrences", index=False)
        # else:
        #     empty_df = pd.DataFrame(columns=["file_name", "Java_file", "Level", "contains_count"])
        #     empty_df.to_excel(writer, sheet_name="FileNameOccurrences", index=False)

    logging.info(f"Excel saved to {METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL} with merged cells for repeated values (order preserved)")

def load_method_line_counts_from_excel(excel_path, sheet_name="Methods"):
    df_methods = pd.read_excel(excel_path, sheet_name=sheet_name, engine="openpyxl")
    df_methods = df_methods.fillna("")

    method_line_map = {}

    for _, row in df_methods.iterrows():
        class_method = str(row["class_method_key"]).strip()
        line_count = row.get("Number_Of_Lines", 0)

        if "." not in class_method:
            continue

        classname, methodname = class_method.split(".", 1)
        try:
            method_line_map[(classname.lower(), methodname.lower())] = int(line_count) if line_count not in (None, "", "nan") and str(line_count).strip() not in ("", "nan", "None") else None
        except (ValueError, TypeError):
            method_line_map[(classname.lower(), methodname.lower())] = None
    return method_line_map

def filter_zero_line_cells(METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL):
    """
    Reads the Excel, blanks out any cell where no_of_lines : 0,
    saves it back, and returns the path.
    """
    import re
    import openpyxl

    zero_pattern = re.compile(r'no_of_lines\s*:\s*0$', re.IGNORECASE)

    wb = openpyxl.load_workbook(METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if zero_pattern.search(cell.value.strip()):
                        cell.value = None

    wb.save(METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL)
    return METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL

# ---------------- MAIN FUNCTION ----------------
def generate_method_level_hierarchy(
    details,
    OUTPUT_DIR,
    excel_path,
    ast_sheet,
    start_file,
    java_folder,
    METHOD_FLOW_OCCURRENCE_DISTRIBUTION,
    DVT_FILES
):
    METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL = os.path.join(
        OUTPUT_DIR, METHOD_FLOW_OCCURRENCE_DISTRIBUTION
    )

    # ---------- Read AST ----------
    df_orig, class_method_map, method_map, graph, filebase_to_nodes, method_to_row = (
        read_ast_and_build_graph(excel_path, sheet_name=ast_sheet)
    )

    methodname_to_rows = build_methodname_lookup(df_orig)
    resolved_cache = {}
    dvt_set = {os.path.splitext(f)[0].lower() for f in DVT_FILES}

    # ✅ LOAD no_of_lines FROM METHODS SHEET
    method_line_map = load_method_line_counts_from_excel(excel_path)

    def drop_row_if_level1_reappears(df_levels):
        """
        Drop entire row if Level_1 value appears anywhere
        in Level_2+ columns across the entire DataFrame.
        """
        if df_levels.empty or df_levels.shape[1] <= 1:
            return df_levels

        level1_col = df_levels.columns[0]
        other_cols = df_levels.columns[1:]

        # Collect all values from Level_2+ globally
        other_values = set(
            str(v).strip()
            for v in df_levels[other_cols].values.ravel()
            if str(v).strip()
        )

        # Identify rows to drop
        rows_to_drop = []
        for idx, val in df_levels[level1_col].items():
            val_str = str(val).strip()
            if val_str and val_str in other_values:
                rows_to_drop.append(idx)

        # Drop rows and reset index
        return df_levels.drop(index=rows_to_drop).reset_index(drop=True)

    all_paths = []
    java_roots = []
    start_file_names = [item.lower() for item in start_file]

    # ---------- Fallback file name column ----------
    if 'file_name_without_ext' not in df_orig.columns:
        df_orig['file_name_without_ext'] = (
            df_orig['filename']
            .apply(lambda x: os.path.splitext(x)[0] if isinstance(x, str) else x)
        )

    # ---------- Start traversal ----------
    for ctrl in start_file_names:
        ctrl_lower = ctrl.lower()

        # Primary match: classname
        df_match = df_orig[df_orig['classname'].str.lower() == ctrl_lower]

        # Fallback: filename
        if df_match.empty:
            df_match = df_orig[
                df_orig['file_name_without_ext'].str.lower() == ctrl_lower
            ]

        start_nodes = (
            df_match[['classname', 'methodname']]
            .drop_duplicates()
            .itertuples(index=False, name=None)
        )

        for classname, methodname in start_nodes:
            paths = expand_lineage_horizontal(
                classname, methodname,
                class_method_map, method_map,
                set(), method_line_map,
                df_orig, method_to_row,
                methodname_to_rows, resolved_cache,
                dvt_set, depth=0
            )

            # ---------- Force root visibility ----------
            root_line_count = method_line_map.get(
                (classname.lower(), methodname.lower()), 0
            )

            root_node = f"{classname}.{methodname} no_of_lines : {root_line_count}"

            if not paths:
                all_paths.append([root_node])
                java_roots.append(classname)
            else:
                all_paths.extend(paths)
                java_roots.extend([classname] * len(paths))

    # ---------- Normalize paths ----------
    cleaned_paths = [[p.split("||")[0] for p in path] for path in all_paths]

    df_levels = build_level_dataframe_from_paths(cleaned_paths)

# ✅ DROP rows whose Level_1 reappears anywhere else
    df_levels = drop_row_if_level1_reappears(df_levels)

    df_inverted = df_levels[df_levels.columns[::-1]]
    
    occ_df = compute_occurrences_from_df_levels(
        df_levels, java_roots, df_orig
    )

    write_all_sheets_pandas(
        METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL,
        df_levels, df_inverted, occ_df
    )
  
    # ✅ Remove cells where no_of_lines is 0
    METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL = filter_zero_line_cells(
        METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL
    )

    return METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL

    return METHOD_FLOW_OCCURRENCE_DISTRIBUTION_EXCEL

