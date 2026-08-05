"""
reorder_method_flow.py
======================
Reorders ONLY the "Original Flow" sheet of
007.2_Method_Detailed_Flow_Occurrence_Distribution.xlsx so that the entire
flow — Level_1 through Level_N — reflects the top-to-bottom method call
sequence as it appears in the codebase.

How ordering works
------------------
Each row in "Original Flow" is a full call path:
    Level_1 → Level_2 → Level_3 → ... → Level_N

Ordering is done in two stages:
  1. Level_1 groups are sorted by the declaration rank of the root method.
  2. Within each Level_1 group, rows are sorted level-by-level using the
     CALL-SITE order: Level_2 children are ranked by the position of their
     invocation inside the Level_1 method body; Level_3 children by their
     invocation position inside that Level_2 body, and so on.

Configuration
-------------
Edit the CONFIG block below, then run:
    python reorder_method_flow.py
"""

import os
import re
import sys
import openpyxl
from openpyxl.styles import Alignment

# ─────────────────────────────────────────────
#  CONFIG  ← edit these paths before running
# ─────────────────────────────────────────────
INPUT_EXCEL  = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution.xlsx"
SRC_ROOT     = r"path/to/java/source/root"
OUTPUT_EXCEL = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution_REORDERED.xlsx"
SRC_EXT      = ".java"    # change to .py / .cs / .kt etc. if needed
SRC_ENCODING = "utf-8"    # fallback to latin-1 handled automatically


# ─────────────────────────────────────────────
#  STEP 1 – scan codebase
#
#  Builds TWO maps:
#
#  decl_rank_map[(class_lower, method_lower)] = int
#      Global declaration rank: order in which the method DECLARATION
#      is first encountered when walking all source files top-to-bottom.
#      Used only for sorting Level_1 groups.
#
#  call_order_map[(caller_class_lower, caller_method_lower)]
#      = [(callee_class_lower, callee_method_lower, call_rank), ...]
#      Records every method CALL SITE inside a method body, in source
#      order.  Used for ordering Level_2+ within their parent group.
# ─────────────────────────────────────────────

# Matches a method/constructor DECLARATION line (must start at column 0
# after optional whitespace and access modifiers).
METHOD_DECL_RE = re.compile(
    r"""
    ^[ \t]*                                         # leading whitespace only
    (?:(?:public|private|protected|static|final
          |abstract|synchronized|native|strictfp)
       \s+)*                                        # zero or more modifiers
    (?:[\w<>\[\].,?\s]+?\s+)?                       # return type (optional for ctors)
    (?P<method>\w+)\s*                              # method / constructor name
    \(                                              # opening paren — declaration
    (?:[^)]*)\)                                     # parameter list
    \s*(?:throws\s+[\w\s,]+)?                       # optional throws
    \s*\{                                           # opening brace = body follows
    """,
    re.VERBOSE | re.MULTILINE,
)

# Matches a method CALL SITE:  identifier( or object.identifier(
# Captures (optional_receiver, method_name).
CALL_SITE_RE = re.compile(
    r"""
    (?:(?P<receiver>[A-Za-z_]\w*)\.)?   # optional receiver (ClassName or var)
    (?P<callee>[A-Za-z_]\w*)            # method name being called
    \s*\(                               # opening paren
    """,
    re.VERBOSE,
)

KEYWORD_SKIP = {
    "if", "else", "while", "for", "switch", "return",
    "try", "catch", "finally", "new", "throw", "assert",
    "class", "interface", "enum", "import", "package",
    "void", "int", "long", "double", "float", "boolean",
    "byte", "char", "short", "String", "static", "final",
    "abstract", "public", "private", "protected", "synchronized",
    "super", "this",
}


def _read_text(path, encoding):
    try:
        with open(path, "r", encoding=encoding) as f:
            return f.read()
    except UnicodeDecodeError:
        with open(path, "r", encoding="latin-1") as f:
            return f.read()


def _strip_comments_and_strings(text):
    """Remove string literals and comments so we don't match inside them."""
    # Remove block comments
    text = re.sub(r'/\*.*?\*/', lambda m: ' ' * len(m.group()), text, flags=re.DOTALL)
    # Remove line comments
    text = re.sub(r'//[^\n]*', '', text)
    # Remove string literals (simple, handles \" escapes)
    text = re.sub(r'"(?:[^"\\]|\\.)*"', '""', text)
    # Remove char literals
    text = re.sub(r"'(?:[^'\\]|\\.)*'", "''", text)
    return text


def build_codebase_order(src_root, ext, encoding):
    """
    Returns (decl_rank_map, call_order_map).

    decl_rank_map  : {(class_lower, method_lower): int}
    call_order_map : {(caller_class_lower, caller_method_lower):
                          [(callee_class_lower_or_None, callee_lower, call_rank)]}
    """
    decl_rank_map  = {}   # (class, method) → declaration rank
    call_order_map = {}   # (caller_class, caller_method) → [callees in order]

    global_decl_rank = 0
    global_call_rank = 0
    file_count = 0

    for dirpath, dirnames, filenames in os.walk(src_root):
        dirnames.sort()
        for fname in sorted(filenames):
            if not fname.endswith(ext):
                continue

            classname = os.path.splitext(fname)[0].lower()
            fpath = os.path.join(dirpath, fname)
            raw_text = _read_text(fpath, encoding)
            text = _strip_comments_and_strings(raw_text)

            # --- collect declaration positions ---
            decl_positions = []   # [(match_start, method_name_lower)]
            for m in METHOD_DECL_RE.finditer(text):
                mname = m.group("method")
                if mname in KEYWORD_SKIP:
                    continue
                mname_lower = mname.lower()
                key = (classname, mname_lower)
                if key not in decl_rank_map:
                    decl_rank_map[key] = global_decl_rank
                    global_decl_rank += 1
                decl_positions.append((m.start(), mname_lower, m.end()))

            # --- for each declared method, find its body and record call sites ---
            # We locate the body by finding the matching closing brace.
            for i, (decl_start, caller_method, body_open_pos) in enumerate(decl_positions):
                # Find opening brace position (METHOD_DECL_RE already requires it)
                brace_pos = text.find('{', body_open_pos - 1)
                if brace_pos == -1:
                    continue

                # Find matching closing brace
                depth = 0
                body_end = brace_pos
                for j in range(brace_pos, len(text)):
                    if text[j] == '{':
                        depth += 1
                    elif text[j] == '}':
                        depth -= 1
                        if depth == 0:
                            body_end = j
                            break

                body = text[brace_pos + 1: body_end]
                caller_key = (classname, caller_method)

                # Scan call sites inside the body in source order
                callees = []
                seen_in_body = set()
                for cm in CALL_SITE_RE.finditer(body):
                    callee_name = cm.group("callee")
                    if callee_name in KEYWORD_SKIP:
                        continue
                    receiver = cm.group("receiver")
                    # Normalise receiver to lower if present
                    receiver_lower = receiver.lower() if receiver else None
                    callee_lower = callee_name.lower()

                    # Deduplicate: only record first call to each unique
                    # (receiver, callee) within this body to preserve order
                    dedup_key = (receiver_lower, callee_lower)
                    if dedup_key not in seen_in_body:
                        seen_in_body.add(dedup_key)
                        callees.append((receiver_lower, callee_lower, global_call_rank))
                        global_call_rank += 1

                if callees:
                    call_order_map[caller_key] = callees

            file_count += 1
            print(f"  [scan] {file_count} files | {global_decl_rank} methods | "
                  f"{global_call_rank} calls — {fname}          ",
                  end="\r", flush=True)

    print(f"\n[codebase scan] Done — {file_count} files, "
          f"{global_decl_rank} declarations, {global_call_rank} call sites.")
    return decl_rank_map, call_order_map


# ─────────────────────────────────────────────
#  STEP 2 – helpers
# ─────────────────────────────────────────────
def extract_classmethod(cell_value):
    """
    'DTBroker3MDB.onMessage no_of_lines : 74' → ('dtbroker3mdb', 'onmessage')
    Returns None if unparseable.
    """
    if not cell_value:
        return None
    base = str(cell_value).split(" ")[0]
    if "." not in base:
        return None
    parts = base.rsplit(".", 1)
    return (parts[0].lower(), parts[1].lower())


def get_decl_rank(cell_value, decl_rank_map):
    """Return declaration rank for a cell value (used for Level_1 sorting)."""
    cm = extract_classmethod(cell_value)
    if cm is None:
        return 10**9
    rank = decl_rank_map.get(cm)
    if rank is not None:
        return rank
    method_lower = cm[1]
    matches = [v for (c, m), v in decl_rank_map.items() if m == method_lower]
    return min(matches) if matches else 10**9


def get_call_rank(parent_cm, child_cell_value, call_order_map):
    """
    Return the call-site rank of child_cell_value as called from parent_cm.
    Falls back to global call rank if parent body not found, then to 10**9.
    parent_cm : (class_lower, method_lower) of the calling method
    """
    child_cm = extract_classmethod(child_cell_value)
    if child_cm is None:
        return 10**9

    callees = call_order_map.get(parent_cm, [])
    child_class, child_method = child_cm

    # 1. Exact match: receiver class + method name
    for (recv, callee, rank) in callees:
        if callee == child_method and (recv is None or recv == child_class):
            return rank

    # 2. Method-name-only match (receiver could be a variable, not class name)
    for (recv, callee, rank) in callees:
        if callee == child_method:
            return rank

    # 3. Not found in parent body — fall back across all callers
    matches = [
        rank
        for entries in call_order_map.values()
        for (_, callee, rank) in entries
        if callee == child_method
    ]
    return min(matches) if matches else 10**9


# ─────────────────────────────────────────────
#  STEP 3 – read Original Flow (expand merges)
# ─────────────────────────────────────────────
def read_original_flow(ws):
    merged_lookup = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_val = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row, col)] = anchor_val

    rows = []
    for r in range(2, ws.max_row + 1):
        row_data = []
        for c in range(1, ws.max_column + 1):
            row_data.append(merged_lookup.get((r, c), ws.cell(r, c).value))
        rows.append(row_data)
    return rows


# ─────────────────────────────────────────────
#  STEP 4 – group rows by Level_1
#  BUG FIX: after merge-expansion every row[0] is non-None, so we
#  must group by VALUE CHANGE, not by None-check.
# ─────────────────────────────────────────────
def group_by_level1(rows):
    groups = []
    current_root = object()   # sentinel — matches nothing initially
    current_rows = []

    for row in rows:
        l1 = row[0]
        if l1 != current_root:          # value changed → new Level_1 group
            if current_rows:
                groups.append((current_root, current_rows))
            current_root = l1
            current_rows = [row]
        else:
            current_rows.append(row)

    if current_rows:
        groups.append((current_root, current_rows))

    return groups


# ─────────────────────────────────────────────
#  STEP 5 – sort
#
#  Level_1 groups → sorted by declaration rank (unchanged, already correct).
#
#  Within each Level_1 group → recursive hierarchical sort:
#    • Find the set of distinct Level_2 values.
#    • Sort those Level_2 values by their CALL-SITE rank inside Level_1.
#    • For each Level_2 value, collect its rows and recurse for Level_3, etc.
#  This guarantees that children at every depth are ordered by the position
#  they are *called* in their parent's body, not by global declaration rank.
# ─────────────────────────────────────────────

def hierarchical_sort(rows, level_idx, decl_rank_map, call_order_map):
    """
    Recursively sort `rows` at the current `level_idx` column.
    Returns a new list of rows in correct call-site order.
    """
    if level_idx >= len(rows[0]):
        return rows

    # Collect unique values at this level, preserving first-occurrence order
    # so we can sub-sort them by call-site rank relative to their parent.
    seen = {}
    for row in rows:
        val = row[level_idx]
        if val not in seen:
            seen[val] = []
        seen[val].append(row)

    if len(seen) == 1:
        # Only one value at this level → recurse deeper
        val = next(iter(seen))
        return hierarchical_sort(rows, level_idx + 1, decl_rank_map, call_order_map)

    # Determine the parent method (level_idx - 1) for call-site rank lookup.
    # All rows share the same ancestors up to level_idx-1 (we're inside one group).
    parent_cm = extract_classmethod(rows[0][level_idx - 1]) if level_idx > 0 else None

    def child_sort_key(child_val):
        if child_val is None:
            return 10**9
        return get_call_rank(parent_cm, child_val, call_order_map)

    sorted_child_vals = sorted(seen.keys(), key=child_sort_key)

    result = []
    for child_val in sorted_child_vals:
        child_rows = seen[child_val]
        # Recurse one level deeper for this child's subtree
        sorted_child_rows = hierarchical_sort(
            child_rows, level_idx + 1, decl_rank_map, call_order_map
        )
        result.extend(sorted_child_rows)

    return result


def sort_all(groups, decl_rank_map, call_order_map):
    # Sort Level_1 groups by declaration rank
    sorted_groups = sorted(groups, key=lambda g: get_decl_rank(g[0], decl_rank_map))

    result = []
    for root_val, group_rows in sorted_groups:
        # Sort within the group starting at Level_2 (index 1),
        # using call-site order relative to each parent.
        sorted_rows = hierarchical_sort(
            group_rows, 1, decl_rank_map, call_order_map
        )
        result.append((root_val, sorted_rows))

    print("\n[reorder] New Level_1 order:")
    for i, (root, _) in enumerate(result, 1):
        r = get_decl_rank(root, decl_rank_map)
        print(f"  {i:>3}. rank={r:<8} {root}")

    return result


# ─────────────────────────────────────────────
#  STEP 6 – write reordered Original Flow sheet
# ─────────────────────────────────────────────
def write_original_flow_sheet(wb, sorted_groups, n_cols):
    if "Original Flow" in wb.sheetnames:
        original_index = wb.sheetnames.index("Original Flow")
        del wb["Original Flow"]
    else:
        original_index = 0

    ws = wb.create_sheet("Original Flow", original_index)

    # Header
    ws.append([f"Level_{i + 1}" for i in range(n_cols)])

    all_rows = []
    for _, group_rows in sorted_groups:
        all_rows.extend(group_rows)

    # Suppress repeated consecutive values per column for merging
    prev = [None] * n_cols
    for row in all_rows:
        padded = (list(row) + [None] * n_cols)[:n_cols]
        out = []
        for col_idx, val in enumerate(padded):
            if val is not None and val == prev[col_idx]:
                out.append(None)
            else:
                out.append(val)
                prev[col_idx] = val
                for deeper in range(col_idx + 1, n_cols):
                    prev[deeper] = None
        ws.append(out)

    # Merge consecutive identical non-None values per column
    merge_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data = list(ws.iter_rows(min_row=2, values_only=True))
    n_rows = len(data)

    for col_idx in range(n_cols):
        r = 0
        while r < n_rows:
            val = data[r][col_idx]
            if val is None:
                r += 1
                continue
            run_end = r + 1
            while run_end < n_rows and data[run_end][col_idx] == val:
                run_end += 1
            if run_end - r > 1:
                ws.merge_cells(
                    start_row=r + 2, start_column=col_idx + 1,
                    end_row=run_end + 1, end_column=col_idx + 1
                )
                ws.cell(r + 2, col_idx + 1).alignment = merge_align
            r = run_end

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    print(f"[write] 'Original Flow' written ({len(all_rows)} data rows, {n_cols} columns)")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"Input  : {INPUT_EXCEL}")
    print(f"Source : {SRC_ROOT}")
    print(f"Output : {OUTPUT_EXCEL}")
    print(f"Ext    : {SRC_EXT}")
    print(f"{'='*60}\n")

    # 1. Build declaration rank map AND call-site order map
    decl_rank_map, call_order_map = build_codebase_order(SRC_ROOT, SRC_EXT, SRC_ENCODING)

    # 2. Load workbook
    print("\n[excel] Loading workbook …")
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws_orig = wb["Original Flow"]
    n_cols = ws_orig.max_column

    # 3. Read Original Flow (merged cells expanded)
    rows = read_original_flow(ws_orig)

    # 4. Group by Level_1 (by value change, not None-check)
    groups = group_by_level1(rows)
    print(f"[group] {len(groups)} Level_1 root groups found.")

    # 5. Sort groups (Level_1 by decl rank; Level_2+ by call-site rank)
    sorted_groups = sort_all(groups, decl_rank_map, call_order_map)

    # 6. Rewrite ONLY 'Original Flow'
    write_original_flow_sheet(wb, sorted_groups, n_cols)

    # 7. Save
    print(f"\n[save] Saving → {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    print("[done] ✅  Reordered workbook saved.\n")


if __name__ == "__main__":
    main()