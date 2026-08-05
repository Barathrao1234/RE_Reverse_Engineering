"""
reorder_method_flow.py
======================
Reorders ONLY the "Original Flow" sheet of
007.2_Method_Detailed_Flow_Occurrence_Distribution.xlsx so that
Level_1 root methods appear in the same order they are encountered when
scanning the Java/source codebase files top-to-bottom.
All other sheets are copied to the output completely untouched.

Configuration
-------------
Edit the CONFIG block below, then run:
    python reorder_method_flow.py
"""

import os
import re
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict

# ─────────────────────────────────────────────
#  CONFIG  ← edit these paths before running
# ─────────────────────────────────────────────
INPUT_EXCEL  = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution.xlsx"
SRC_ROOT     = r"path/to/java/source/root"
OUTPUT_EXCEL = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution_REORDERED.xlsx"
SRC_EXT      = ".java"    # change to .py / .cs / .kt etc. if needed
SRC_ENCODING = "utf-8"    # fallback to latin-1 is handled automatically


# ─────────────────────────────────────────────
#  STEP 1 – build codebase method ordering
# ─────────────────────────────────────────────
METHOD_DECL_RE = re.compile(
    r"""
    ^[ \t]*
    (?:public|private|protected|static|final|abstract|synchronized|native|strictfp|\s)*
    [\w<>\[\].,?\s]+?       # return type
    \b(?P<method>\w+)\s*\(  # method name followed by (
    """,
    re.VERBOSE | re.MULTILINE
)

def _read_text(path, encoding):
    try:
        with open(path, "r", encoding=encoding) as f:
            return f.read()
    except UnicodeDecodeError:
        with open(path, "r", encoding="latin-1") as f:
            return f.read()


def build_codebase_order(src_root, ext, encoding):
    """
    Walk src_root, find every method declaration and record
    (classname_lower, methodname_lower) → global_rank (lower = earlier).
    Files are visited in sorted order so the result is deterministic.
    """
    rank_map = {}
    rank = 0

    for dirpath, dirnames, filenames in os.walk(src_root):
        dirnames.sort()
        for fname in sorted(filenames):
            if not fname.endswith(ext):
                continue
            classname = os.path.splitext(fname)[0]
            fpath = os.path.join(dirpath, fname)
            text = _read_text(fpath, encoding)

            for m in METHOD_DECL_RE.finditer(text):
                mname = m.group("method")
                if mname in {"if", "else", "while", "for", "switch", "return",
                             "try", "catch", "finally", "new", "throw", "assert",
                             "class", "interface", "enum", "import", "package",
                             "void", "int", "long", "double", "float", "boolean",
                             "byte", "char", "short", "String", "static", "final",
                             "abstract", "public", "private", "protected", "synchronized"}:
                    continue
                key = (classname.lower(), mname.lower())
                if key not in rank_map:
                    rank_map[key] = rank
                    rank += 1

    print(f"[codebase scan] {rank} unique (class, method) pairs found in {src_root}")
    return rank_map


# ─────────────────────────────────────────────
#  STEP 2 – read Original Flow, expand merged cells
# ─────────────────────────────────────────────
def extract_classmethod(cell_value):
    """
    'DTBroker3MDB.onMessage no_of_lines : 74'  →  ('dtbroker3mdb', 'onmessage')
    Returns None if the string doesn't look like ClassName.method...
    """
    if not cell_value:
        return None
    base = cell_value.split(" ")[0]
    if "." not in base:
        return None
    parts = base.rsplit(".", 1)
    return (parts[0].lower(), parts[1].lower())


def read_original_flow(ws):
    """
    Read 'Original Flow' sheet respecting merged cells.
    Returns a list of rows where each row is a list of cell values (str or None).
    Merged cells are expanded: the top-left anchor value is repeated into every cell.
    """
    merged_lookup = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_val = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row, col)] = anchor_val

    rows = []
    for r in range(2, ws.max_row + 1):   # skip header row 1
        row_data = []
        for c in range(1, ws.max_column + 1):
            if (r, c) in merged_lookup:
                row_data.append(merged_lookup[(r, c)])
            else:
                row_data.append(ws.cell(r, c).value)
        rows.append(row_data)
    return rows


# ─────────────────────────────────────────────
#  STEP 3 – group rows by Level_1 root
# ─────────────────────────────────────────────
def group_by_level1(rows):
    """
    Split the flat row list into groups, one per Level_1 root entry.
    Each group is (root_value, [rows]).
    """
    groups = []
    current_root = None
    current_rows = []

    for row in rows:
        l1 = row[0]
        if l1 is not None:
            if current_root is not None:
                groups.append((current_root, current_rows))
            current_root = l1
            current_rows = [row]
        else:
            current_rows.append(row)

    if current_root is not None:
        groups.append((current_root, current_rows))

    return groups


# ─────────────────────────────────────────────
#  STEP 4 – sort groups by codebase rank
# ─────────────────────────────────────────────
def sort_groups(groups, rank_map):
    def sort_key(g):
        cm = extract_classmethod(g[0])
        if cm is None:
            return (10**9, groups.index(g))
        rank = rank_map.get(cm)
        if rank is None:
            # Fall back: match on method name alone across all classes
            method_lower = cm[1]
            matches = [v for (c, m), v in rank_map.items() if m == method_lower]
            rank = min(matches) if matches else 10**9
        return (rank, 0)

    sorted_groups = sorted(groups, key=sort_key)

    print("\n[reorder] New Level_1 order:")
    for i, (root, _) in enumerate(sorted_groups, 1):
        cm = extract_classmethod(root)
        r = rank_map.get(cm, "?") if cm else "?"
        print(f"  {i:>3}. rank={r:<8} {root}")

    return sorted_groups


# ─────────────────────────────────────────────
#  STEP 5 – write reordered Original Flow sheet
# ─────────────────────────────────────────────
def write_original_flow_sheet(wb, sorted_rows, n_cols):
    """
    Delete and recreate 'Original Flow' with the reordered rows and merged cells.
    All other sheets in wb are left completely untouched.
    """
    if "Original Flow" in wb.sheetnames:
        # Remember the tab position so we can restore it
        original_index = wb.sheetnames.index("Original Flow")
        del wb["Original Flow"]
    else:
        original_index = 0

    ws = wb.create_sheet("Original Flow", original_index)

    # Write header
    headers = [f"Level_{i + 1}" for i in range(n_cols)]
    ws.append(headers)

    # Write data rows — suppress repeated Level_1 values (will be merged below)
    prev_l1 = None
    for row in sorted_rows:
        padded = row + [None] * (n_cols - len(row))
        if padded[0] == prev_l1:
            padded[0] = None
        elif padded[0] is not None:
            prev_l1 = padded[0]
        ws.append(padded)

    # Merge consecutive identical non-None values in every column
    merge_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data = list(ws.iter_rows(min_row=2, values_only=True))
    n_rows = len(data)

    for col_idx in range(n_cols):
        r = 0
        while r < n_rows:
            val = data[r][col_idx]
            if val is None:
                r += 1
                continue
            run_end = r + 1
            while run_end < n_rows and data[run_end][col_idx] == val:
                run_end += 1
            if run_end - r > 1:
                ws.merge_cells(
                    start_row=r + 2, start_column=col_idx + 1,
                    end_row=run_end + 1, end_column=col_idx + 1
                )
                ws.cell(r + 2, col_idx + 1).alignment = merge_align
            r = run_end

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    print(f"[write] 'Original Flow' written ({len(sorted_rows)} data rows, {n_cols} columns)")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"Input  : {INPUT_EXCEL}")
    print(f"Source : {SRC_ROOT}")
    print(f"Output : {OUTPUT_EXCEL}")
    print(f"Ext    : {SRC_EXT}")
    print(f"{'='*60}\n")

    # 1. Build codebase ordering
    rank_map = build_codebase_order(SRC_ROOT, SRC_EXT, SRC_ENCODING)

    # 2. Load workbook (all sheets intact)
    print("\n[excel] Loading workbook …")
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws_orig = wb["Original Flow"]
    n_cols = ws_orig.max_column

    # 3. Read Original Flow rows (merged cells expanded)
    rows = read_original_flow(ws_orig)

    # 4. Group by Level_1
    groups = group_by_level1(rows)
    print(f"[group] {len(groups)} Level_1 root groups found.")

    # 5. Sort groups by codebase rank
    sorted_groups = sort_groups(groups, rank_map)

    # 6. Flatten back to a single row list
    sorted_rows = []
    for _, group_rows in sorted_groups:
        sorted_rows.extend(group_rows)

    # 7. Rewrite ONLY 'Original Flow' — all other sheets untouched
    write_original_flow_sheet(wb, sorted_rows, n_cols)

    # 8. Save
    print(f"\n[save] Saving → {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    print("[done] ✅  Reordered workbook saved.\n")


if __name__ == "__main__":
    main()
