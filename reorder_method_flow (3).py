"""
reorder_method_flow.py
======================
Reorders ONLY the "Original Flow" sheet of
007.2_Method_Detailed_Flow_Occurrence_Distribution.xlsx so that the entire
flow — Level_1 through Level_N — reflects the top-to-bottom method call
sequence as it appears in the codebase.

How ordering works
------------------
Each row in "Original Flow" is a full call path:
    Level_1 → Level_2 → Level_3 → ... → Level_N

Ordering is done in two stages:
  1. Level_1 groups are sorted by the line rank of the root method in
     the codebase (which file / line it is declared in).
  2. Within each Level_1 group, rows (paths) are sorted left-to-right
     by the codebase rank of each successive level — so Level_2 children
     appear in the order they are called inside the Level_1 method body,
     Level_3 children in the order called inside that Level_2 method, etc.

Configuration
-------------
Edit the CONFIG block below, then run:
    python reorder_method_flow.py
"""

import os
import re
import sys
import openpyxl
from openpyxl.styles import Alignment

# ─────────────────────────────────────────────
#  CONFIG  ← edit these paths before running
# ─────────────────────────────────────────────
INPUT_EXCEL  = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution.xlsx"
SRC_ROOT     = r"path/to/java/source/root"
OUTPUT_EXCEL = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution_REORDERED.xlsx"
SRC_EXT      = ".java"    # change to .py / .cs / .kt etc. if needed
SRC_ENCODING = "utf-8"    # fallback to latin-1 handled automatically


# ─────────────────────────────────────────────
#  STEP 1 – scan codebase, build rank map
#  rank_map[(classname_lower, methodname_lower)] = global_rank
#  rank = order in which the method declaration is first encountered
#         when walking all source files top-to-bottom
# ─────────────────────────────────────────────
METHOD_DECL_RE = re.compile(
    r"""
    ^[ \t]*
    (?:public|private|protected|static|final|abstract|synchronized|native|strictfp|\s)*
    [\w<>\[\].,?\s]+?
    \b(?P<method>\w+)\s*\(
    """,
    re.VERBOSE | re.MULTILINE
)

KEYWORD_SKIP = {
    "if", "else", "while", "for", "switch", "return",
    "try", "catch", "finally", "new", "throw", "assert",
    "class", "interface", "enum", "import", "package",
    "void", "int", "long", "double", "float", "boolean",
    "byte", "char", "short", "String", "static", "final",
    "abstract", "public", "private", "protected", "synchronized"
}

def _read_text(path, encoding):
    try:
        with open(path, "r", encoding=encoding) as f:
            return f.read()
    except UnicodeDecodeError:
        with open(path, "r", encoding="latin-1") as f:
            return f.read()


def build_codebase_order(src_root, ext, encoding):
    rank_map = {}
    rank = 0
    file_count = 0

    for dirpath, dirnames, filenames in os.walk(src_root):
        dirnames.sort()
        for fname in sorted(filenames):
            if not fname.endswith(ext):
                continue
            classname = os.path.splitext(fname)[0]
            fpath = os.path.join(dirpath, fname)
            text = _read_text(fpath, encoding)

            for m in METHOD_DECL_RE.finditer(text):
                mname = m.group("method")
                if mname in KEYWORD_SKIP:
                    continue
                key = (classname.lower(), mname.lower())
                if key not in rank_map:
                    rank_map[key] = rank
                    rank += 1

            file_count += 1
            print(f"  [scan] {file_count} files | {rank} methods parsed — {fname}          ",
                  end="\r", flush=True)

    print(f"\n[codebase scan] Done — {file_count} files, {rank} unique methods found.")
    return rank_map


# ─────────────────────────────────────────────
#  STEP 2 – helpers
# ─────────────────────────────────────────────
def extract_classmethod(cell_value):
    """
    'DTBroker3MDB.onMessage no_of_lines : 74' → ('dtbroker3mdb', 'onmessage')
    Returns None if unparseable.
    """
    if not cell_value:
        return None
    base = cell_value.split(" ")[0]
    if "." not in base:
        return None
    parts = base.rsplit(".", 1)
    return (parts[0].lower(), parts[1].lower())


def get_rank(cell_value, rank_map):
    """Return codebase rank for a cell value, or a large fallback."""
    cm = extract_classmethod(cell_value)
    if cm is None:
        return 10**9
    rank = rank_map.get(cm)
    if rank is not None:
        return rank
    # Fallback: match method name only across all classes
    method_lower = cm[1]
    matches = [v for (c, m), v in rank_map.items() if m == method_lower]
    return min(matches) if matches else 10**9


# ─────────────────────────────────────────────
#  STEP 3 – read Original Flow (expand merges)
# ─────────────────────────────────────────────
def read_original_flow(ws):
    merged_lookup = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_val = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row, col)] = anchor_val

    rows = []
    for r in range(2, ws.max_row + 1):
        row_data = []
        for c in range(1, ws.max_column + 1):
            row_data.append(merged_lookup.get((r, c), ws.cell(r, c).value))
        rows.append(row_data)
    return rows


# ─────────────────────────────────────────────
#  STEP 4 – group rows by Level_1
# ─────────────────────────────────────────────
def group_by_level1(rows):
    groups = []
    current_root = None
    current_rows = []

    for row in rows:
        l1 = row[0]
        if l1 is not None:
            if current_root is not None:
                groups.append((current_root, current_rows))
            current_root = l1
            current_rows = [row]
        else:
            current_rows.append(row)

    if current_root is not None:
        groups.append((current_root, current_rows))

    return groups


# ─────────────────────────────────────────────
#  STEP 5 – sort: Level_1 groups AND within
#            each group sort paths by level ranks
# ─────────────────────────────────────────────
def path_sort_key(row, rank_map):
    """
    Sort key for a single path-row.
    Compares level by level left-to-right using codebase rank.
    None cells (empty levels) sort after non-None so deeper paths
    for the same prefix stay grouped together.
    """
    return tuple(
        get_rank(cell, rank_map) if cell else 10**9
        for cell in row
    )


def sort_all(groups, rank_map):
    # Sort Level_1 groups by their root rank
    def l1_key(g):
        return get_rank(g[0], rank_map)

    sorted_groups = sorted(groups, key=l1_key)

    # Within each group, sort rows by full path rank (level by level)
    result = []
    for root_val, group_rows in sorted_groups:
        sorted_rows = sorted(group_rows, key=lambda row: path_sort_key(row, rank_map))
        result.append((root_val, sorted_rows))

    # Print new Level_1 order
    print("\n[reorder] New Level_1 order:")
    for i, (root, _) in enumerate(result, 1):
        r = get_rank(root, rank_map)
        print(f"  {i:>3}. rank={r:<8} {root}")

    return result


# ─────────────────────────────────────────────
#  STEP 6 – write reordered Original Flow sheet
# ─────────────────────────────────────────────
def write_original_flow_sheet(wb, sorted_groups, n_cols):
    if "Original Flow" in wb.sheetnames:
        original_index = wb.sheetnames.index("Original Flow")
        del wb["Original Flow"]
    else:
        original_index = 0

    ws = wb.create_sheet("Original Flow", original_index)

    # Header
    ws.append([f"Level_{i + 1}" for i in range(n_cols)])

    # Flatten all groups into rows, suppressing repeated consecutive values
    # per column so merging works correctly
    all_rows = []
    for _, group_rows in sorted_groups:
        all_rows.extend(group_rows)

    # Track last written value per column to suppress duplicates for merging
    prev = [None] * n_cols
    for row in all_rows:
        padded = (row + [None] * n_cols)[:n_cols]
        out = []
        for col_idx, val in enumerate(padded):
            if val is not None and val == prev[col_idx]:
                out.append(None)   # suppress — will be covered by merge
            else:
                out.append(val)
                prev[col_idx] = val
                # Reset all deeper columns when a shallower one changes
                for deeper in range(col_idx + 1, n_cols):
                    prev[deeper] = None
        ws.append(out)

    # Merge consecutive identical non-None values per column
    merge_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data = list(ws.iter_rows(min_row=2, values_only=True))
    n_rows = len(data)

    for col_idx in range(n_cols):
        r = 0
        while r < n_rows:
            val = data[r][col_idx]
            if val is None:
                r += 1
                continue
            run_end = r + 1
            while run_end < n_rows and data[run_end][col_idx] == val:
                run_end += 1
            if run_end - r > 1:
                ws.merge_cells(
                    start_row=r + 2, start_column=col_idx + 1,
                    end_row=run_end + 1, end_column=col_idx + 1
                )
                ws.cell(r + 2, col_idx + 1).alignment = merge_align
            r = run_end

    # Auto-width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    print(f"[write] 'Original Flow' written ({len(all_rows)} data rows, {n_cols} columns)")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"Input  : {INPUT_EXCEL}")
    print(f"Source : {SRC_ROOT}")
    print(f"Output : {OUTPUT_EXCEL}")
    print(f"Ext    : {SRC_EXT}")
    print(f"{'='*60}\n")

    # 1. Build codebase rank map
    rank_map = build_codebase_order(SRC_ROOT, SRC_EXT, SRC_ENCODING)

    # 2. Load workbook
    print("\n[excel] Loading workbook …")
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws_orig = wb["Original Flow"]
    n_cols = ws_orig.max_column

    # 3. Read Original Flow (merged cells expanded)
    rows = read_original_flow(ws_orig)

    # 4. Group by Level_1
    groups = group_by_level1(rows)
    print(f"[group] {len(groups)} Level_1 root groups found.")

    # 5. Sort Level_1 groups AND sort rows within each group by full path rank
    sorted_groups = sort_all(groups, rank_map)

    # 6. Rewrite ONLY 'Original Flow' — all other sheets untouched
    write_original_flow_sheet(wb, sorted_groups, n_cols)

    # 7. Save
    print(f"\n[save] Saving → {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    print("[done] ✅  Reordered workbook saved.\n")


if __name__ == "__main__":
    main()
