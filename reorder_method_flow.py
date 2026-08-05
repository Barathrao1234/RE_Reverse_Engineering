"""
reorder_method_flow.py
======================
Reorders 007.2_Method_Detailed_Flow_Occurrence_Distribution.xlsx so that
Level_1 root methods appear in the same order they are encountered when
scanning the Java/source codebase files top-to-bottom.

Configuration
-------------
Edit the CONFIG block below, then run:
    python reorder_method_flow.py

What it does
------------
1. Reads the existing Excel (all sheets, all formatting/merges preserved via openpyxl).
2. Walks the source tree and records the line-number of each method declaration,
   producing a global ordering: (file_path, line_number) → rank.
3. Extracts each Level_1 root name (e.g. "DTBroker3MDB.onMessage no_of_lines : 74"),
   strips to "ClassName.methodName", and looks it up in the codebase ranking.
4. Sorts the Level_1 groups by that rank; unrecognised roots keep their original
   relative order at the end.
5. Rebuilds "Original Flow" and "Inverted Flow" with merged cells, then rewrites
   every dependent sheet (File_Occurrences, single_call, FileNames_*, FullRows_*)
   by re-running the same aggregation logic on the reordered data.
6. Saves the output without touching any sheet that has no dependency on the order.
"""

import os
import re
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict

# ─────────────────────────────────────────────
#  CONFIG  ← edit these paths before running
# ─────────────────────────────────────────────
INPUT_EXCEL  = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution.xlsx"
SRC_ROOT     = r"path/to/java/source/root"
OUTPUT_EXCEL = r"path/to/007_2_Method_Detailed_Flow_Occurrence_Distribution_REORDERED.xlsx"
SRC_EXT      = ".java"    # change to .py / .cs / .kt etc. if needed
SRC_ENCODING = "utf-8"    # fallback to latin-1 is handled automatically


# ─────────────────────────────────────────────
#  STEP 2 – build codebase method ordering
# ─────────────────────────────────────────────
METHOD_DECL_RE = re.compile(
    r"""
    ^[ \t]*
    (?:public|private|protected|static|final|abstract|synchronized|native|strictfp|\s)*
    [\w<>\[\].,?\s]+?       # return type
    \b(?P<method>\w+)\s*\(  # method name followed by (
    """,
    re.VERBOSE | re.MULTILINE
)

def _read_text(path, encoding):
    try:
        with open(path, "r", encoding=encoding) as f:
            return f.read()
    except UnicodeDecodeError:
        with open(path, "r", encoding="latin-1") as f:
            return f.read()


def build_codebase_order(src_root, ext, encoding):
    """
    Walk src_root, find every method declaration and record
    (classname_lower, methodname_lower) → global_rank  (lower = earlier).
    Files are visited in sorted order so the result is deterministic.
    """
    # (classname_lower, methodname_lower) -> rank (first occurrence wins)
    rank_map = {}
    rank = 0

    for dirpath, dirnames, filenames in os.walk(src_root):
        dirnames.sort()
        for fname in sorted(filenames):
            if not fname.endswith(ext):
                continue
            classname = os.path.splitext(fname)[0]
            fpath = os.path.join(dirpath, fname)
            text = _read_text(fpath, encoding)

            # Find CLASS declaration to anchor the classname correctly.
            # We trust the file name == class name (standard Java convention).
            # Then record every method declaration in file order.
            for m in METHOD_DECL_RE.finditer(text):
                mname = m.group("method")
                # Skip obvious keywords that slip through the regex
                if mname in {"if", "else", "while", "for", "switch", "return",
                             "try", "catch", "finally", "new", "throw", "assert",
                             "class", "interface", "enum", "import", "package",
                             "void", "int", "long", "double", "float", "boolean",
                             "byte", "char", "short", "String", "static", "final",
                             "abstract", "public", "private", "protected", "synchronized"}:
                    continue
                key = (classname.lower(), mname.lower())
                if key not in rank_map:
                    rank_map[key] = rank
                    rank += 1

    print(f"[codebase scan] {rank} unique (class, method) pairs found in {src_root}")
    return rank_map


# ─────────────────────────────────────────────
#  STEP 3 – read existing Excel, group by Level_1
# ─────────────────────────────────────────────
def extract_classmethod(cell_value):
    """
    'DTBroker3MDB.onMessage no_of_lines : 74'  →  ('dtbroker3mdb', 'onmessage')
    Returns None if the string doesn't look like ClassName.method...
    """
    if not cell_value:
        return None
    # Strip everything after the first space following the method name
    base = cell_value.split(" ")[0]   # e.g. 'DTBroker3MDB.onMessage'
    if "." not in base:
        return None
    parts = base.rsplit(".", 1)
    return (parts[0].lower(), parts[1].lower())


def read_original_flow(ws):
    """
    Read 'Original Flow' sheet respecting merged cells.
    Returns a list of rows, where each row is a list of cell values (str or None).
    Merged cells are expanded: the top-left anchor value is repeated.
    """
    # Build a value grid that fills in merged-cell spans
    merged_lookup = {}   # (row, col) -> value from anchor
    for merged_range in ws.merged_cells.ranges:
        anchor_val = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row, col)] = anchor_val

    rows = []
    for r in range(2, ws.max_row + 1):   # skip header row 1
        row_data = []
        for c in range(1, ws.max_column + 1):
            if (r, c) in merged_lookup:
                row_data.append(merged_lookup[(r, c)])
            else:
                row_data.append(ws.cell(r, c).value)
        rows.append(row_data)
    return rows


def group_by_level1(rows):
    """
    Split the flat row list into groups, one per Level_1 root entry.
    Each group is (root_value, [rows]).
    """
    groups = []
    current_root = None
    current_rows = []

    for row in rows:
        l1 = row[0]
        if l1 is not None:   # new root
            if current_root is not None:
                groups.append((current_root, current_rows))
            current_root = l1
            current_rows = [row]
        else:
            current_rows.append(row)

    if current_root is not None:
        groups.append((current_root, current_rows))

    return groups


# ─────────────────────────────────────────────
#  STEP 4 – sort groups by codebase rank
# ─────────────────────────────────────────────
def sort_groups(groups, rank_map):
    def sort_key(g):
        root_val = g[0]
        cm = extract_classmethod(root_val)
        if cm is None:
            return (10**9, groups.index(g))   # unrecognised → end
        rank = rank_map.get(cm)
        if rank is None:
            # Try just the method name across all classes
            method_lower = cm[1]
            matches = [v for (c, m), v in rank_map.items() if m == method_lower]
            rank = min(matches) if matches else 10**9
        return (rank, 0)

    sorted_groups = sorted(groups, key=sort_key)

    # Report
    print("\n[reorder] New Level_1 order:")
    for i, (root, _) in enumerate(sorted_groups, 1):
        cm = extract_classmethod(root)
        r = rank_map.get(cm, "?") if cm else "?"
        print(f"  {i:>3}. rank={r:<8} {root}")

    return sorted_groups


# ─────────────────────────────────────────────
#  STEP 5 – rebuild Original Flow sheet with merges
# ─────────────────────────────────────────────
def write_flow_sheet(wb, sheet_name, rows_data, invert=False):
    """
    Write rows_data to sheet_name with merged cells for repeated consecutive values.
    rows_data: list of row lists (already in final order, Level_1 expanded).
    invert: if True, reverse columns (for 'Inverted Flow').
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    n_cols = max((len(r) for r in rows_data), default=0)

    # Build headers
    if invert:
        headers = [f"Level_{n_cols - i}" for i in range(n_cols)]
    else:
        headers = [f"Level_{i + 1}" for i in range(n_cols)]

    ws.append(headers)

    # Write data rows, deduplicate Level_1 (write None where same as previous)
    prev_l1 = None
    for row in rows_data:
        padded = row + [None] * (n_cols - len(row))
        if invert:
            padded = list(reversed(padded))
        # Deduplicate Level_1 for display (will be merged later)
        if not invert:
            if padded[0] == prev_l1:
                padded[0] = None
            elif padded[0] is not None:
                prev_l1 = padded[0]
        ws.append(padded)

    # Apply merged cells for consecutive identical non-None values per column
    merge_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data = list(ws.iter_rows(min_row=2, values_only=True))
    n_rows = len(data)

    for col_idx in range(n_cols):
        col_letter = ws.cell(1, col_idx + 1).column_letter
        r = 0
        while r < n_rows:
            val = data[r][col_idx]
            if val is None:
                r += 1
                continue
            # Find run length
            run_end = r + 1
            while run_end < n_rows and data[run_end][col_idx] == val:
                run_end += 1
            if run_end - r > 1:
                start_excel = r + 2    # +1 for header, +1 for 1-based
                end_excel   = run_end + 1
                ws.merge_cells(
                    start_row=start_excel, start_column=col_idx + 1,
                    end_row=end_excel,     end_column=col_idx + 1
                )
                ws.cell(start_excel, col_idx + 1).alignment = merge_align
            r = run_end

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                v = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    return ws


# ─────────────────────────────────────────────
#  STEP 6 – rebuild dependent sheets
# ─────────────────────────────────────────────
def recompute_file_occurrences(sorted_rows, n_cols):
    """
    Recompute File_Occurrences: for each unique file_name (ClassName.method extracted
    from any cell), count in which Level columns it appears.
    Returns list of dicts with keys: file_name, Level_1..Level_n
    """
    # Collect all unique file_name values (ClassName.method) across all cells
    file_level_counts = defaultdict(lambda: defaultdict(int))

    for row in sorted_rows:
        for col_idx, cell in enumerate(row):
            if cell:
                base = cell.split(" ")[0]  # 'ClassName.method'
                if "." in base:
                    level_key = f"Level_{col_idx + 1}"
                    file_level_counts[base][level_key] += 1

    level_cols = [f"Level_{i+1}" for i in range(n_cols)]
    records = []
    for fname in sorted(file_level_counts.keys()):
        rec = {"file_name": fname}
        for lc in level_cols:
            rec[lc] = file_level_counts[fname].get(lc, 0)
        records.append(rec)
    return records


def recompute_single_call(sorted_rows, n_cols):
    """
    single_call: rows where a file_name appears in exactly one level and exactly once total.
    """
    file_level_counts = defaultdict(lambda: defaultdict(int))

    for row in sorted_rows:
        for col_idx, cell in enumerate(row):
            if cell:
                base = cell.split(" ")[0]
                if "." in base:
                    level_key = f"level_{col_idx + 1}"
                    file_level_counts[base][level_key] += 1

    level_cols = [f"level_{i+1}" for i in range(n_cols)]
    records = []
    for fname in sorted(file_level_counts.keys()):
        counts = file_level_counts[fname]
        total = sum(counts.values())
        non_zero = sum(1 for v in counts.values() if v > 0)
        if total == 1 and non_zero == 1:
            rec = {"file_name": fname}
            for lc in level_cols:
                rec[lc] = counts.get(lc, 0)
            records.append(rec)
    return records


def categorise_by_line_count(sorted_rows, n_cols):
    """
    Categorise file_names by their no_of_lines value into buckets:
    GT500, 200_500, 100_200, 50_100, LT50.
    Returns dict bucket_name -> list of file_names (sorted).
    """
    # Extract (file_name, line_count) from Level_1 column only
    file_lines = {}
    for row in sorted_rows:
        for cell in row:
            if cell:
                m = re.match(r'(\S+)\s+no_of_lines\s*:\s*(\d+)', cell)
                if m:
                    fname = m.group(1)
                    lines = int(m.group(2))
                    if fname not in file_lines:
                        file_lines[fname] = lines

    buckets = {
        "GT500":    [],
        "200_500":  [],
        "100_200":  [],
        "50_100":   [],
        "LT50":     [],
    }
    for fname, lines in file_lines.items():
        if lines > 500:
            buckets["GT500"].append(fname)
        elif lines >= 200:
            buckets["200_500"].append(fname)
        elif lines >= 100:
            buckets["100_200"].append(fname)
        elif lines >= 50:
            buckets["50_100"].append(fname)
        else:
            buckets["LT50"].append(fname)

    for k in buckets:
        buckets[k] = sorted(buckets[k])
    return buckets


def write_dependent_sheets(wb, sorted_rows, n_cols, original_wb):
    """
    Rebuild all derived sheets using the reordered sorted_rows.
    Sheets not related to ordering are copied unchanged from original_wb.
    """
    level_cols_lower = [f"level_{i+1}" for i in range(n_cols)]
    level_cols_upper = [f"Level_{i+1}" for i in range(n_cols)]

    # ── File_Occurrences ─────────────────────
    occ_records = recompute_file_occurrences(sorted_rows, n_cols)
    _write_simple_sheet(wb, "File_Occurrences",
                        ["file_name"] + level_cols_upper, occ_records)

    # ── single_call ───────────────────────────
    sc_records = recompute_single_call(sorted_rows, n_cols)
    _write_simple_sheet(wb, "single_call",
                        ["file_name"] + level_cols_lower, sc_records)

    # ── Bucket sheets ─────────────────────────
    buckets = categorise_by_line_count(sorted_rows, n_cols)
    # Build a lookup: file_name -> full row record (level counts)
    occ_lookup = {r["file_name"]: r for r in occ_records}

    bucket_map = {
        "GT500":   ("FileNames_GT500",   "FullRows_GT500"),
        "200_500": ("FileNames_200_500", "FullRows_200_500"),
        "100_200": ("FileNames_100_200", "FullRows_100_200"),
        "50_100":  ("FileNames_50_100",  "FullRows_50_100"),
        "LT50":    ("FileNames_LT50",    "FullRows_LT50"),
    }

    for bucket_key, (names_sheet, rows_sheet) in bucket_map.items():
        fnames = buckets[bucket_key]

        # FileNames_* sheet: single column
        _write_simple_sheet(wb, names_sheet,
                            ["file_name"],
                            [{"file_name": f} for f in fnames])

        # FullRows_* sheet: file_name + level counts (lowercase headers)
        full_records = []
        for f in fnames:
            rec = {"file_name": f}
            base_rec = occ_lookup.get(f, {})
            for lc_upper, lc_lower in zip(level_cols_upper, level_cols_lower):
                rec[lc_lower] = base_rec.get(lc_upper, 0)
            full_records.append(rec)
        _write_simple_sheet(wb, rows_sheet,
                            ["file_name"] + level_cols_lower,
                            full_records)


def _write_simple_sheet(wb, sheet_name, headers, records):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for rec in records:
        ws.append([rec.get(h, 0) for h in headers])
    # Auto-width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)


# ─────────────────────────────────────────────
#  STEP 7 – preserve sheet order matching original
# ─────────────────────────────────────────────
DESIRED_SHEET_ORDER = [
    "Original Flow", "Inverted Flow",
    "File_Occurrences", "single_call",
    "FileNames_GT500", "FullRows_GT500",
    "FileNames_200_500", "FullRows_200_500",
    "FileNames_100_200", "FullRows_100_200",
    "FileNames_50_100", "FullRows_50_100",
    "FileNames_LT50", "FullRows_LT50",
]

def reorder_sheets(wb):
    existing = wb.sheetnames
    ordered = [s for s in DESIRED_SHEET_ORDER if s in existing]
    others  = [s for s in existing if s not in ordered]
    final   = ordered + others
    wb._sheets.sort(key=lambda ws: final.index(ws.title) if ws.title in final else 999)


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"Input  : {INPUT_EXCEL}")
    print(f"Source : {SRC_ROOT}")
    print(f"Output : {OUTPUT_EXCEL}")
    print(f"Ext    : {SRC_EXT}")
    print(f"{'='*60}\n")

    # 1. Build codebase ordering
    rank_map = build_codebase_order(SRC_ROOT, SRC_EXT, SRC_ENCODING)

    # 2. Load existing workbook
    print("\n[excel] Loading workbook …")
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws_orig = wb["Original Flow"]

    # 3. Read rows (merged cells expanded)
    rows = read_original_flow(ws_orig)
    n_cols = ws_orig.max_column

    # 4. Group by Level_1
    groups = group_by_level1(rows)
    print(f"\n[group] {len(groups)} Level_1 root groups found.")

    # 5. Sort groups
    sorted_groups = sort_groups(groups, rank_map)

    # 6. Flatten back to rows list
    sorted_rows = []
    for root_val, group_rows in sorted_groups:
        sorted_rows.extend(group_rows)

    # 7. Write 'Original Flow'
    print("\n[write] Writing 'Original Flow' …")
    write_flow_sheet(wb, "Original Flow", sorted_rows, invert=False)

    # 8. Write 'Inverted Flow'
    print("[write] Writing 'Inverted Flow' …")
    write_flow_sheet(wb, "Inverted Flow", sorted_rows, invert=True)

    # 9. Rebuild dependent sheets
    print("[write] Rebuilding dependent sheets …")
    write_dependent_sheets(wb, sorted_rows, n_cols, wb)

    # 10. Fix sheet order
    reorder_sheets(wb)

    # 11. Save
    print(f"\n[save] Saving → {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    print("[done] ✅  Reordered workbook saved.\n")


if __name__ == "__main__":
    main()
